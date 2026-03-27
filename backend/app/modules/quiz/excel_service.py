"""Excel import/export for quiz questions.

Handles bulk question management via Excel files with Arabic headers.
"""

import uuid
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.enums import QuestionDifficulty, QuestionStatus, QuestionType
from app.modules.quiz.models import Question

# ── Column mapping (Arabic headers) ─────────────────────────────────────

EXPECTED_HEADERS = [
    "السؤال",
    "النوع",
    "الخيار_1",
    "الخيار_2",
    "الخيار_3",
    "الخيار_4",
    "الإجابة_الصحيحة",
    "النقاط",
    "الصعوبة",
    "الفئة",
]

VALID_TYPES = {
    "multiple_choice": QuestionType.MULTIPLE_CHOICE,
    "true_false": QuestionType.TRUE_FALSE,
}

VALID_DIFFICULTIES = {
    "easy": QuestionDifficulty.EASY,
    "medium": QuestionDifficulty.MEDIUM,
    "hard": QuestionDifficulty.HARD,
}


# ── Import ───────────────────────────────────────────────────────────────


async def import_questions_from_excel(
    session: AsyncSession,
    file_bytes: bytes,
    group_id: uuid.UUID,
) -> dict:
    """Parse Excel file and create Question records in the group.

    Returns: {
        "total_rows": int,
        "imported": int,
        "errors": [{"row": int, "error": str}],
    }
    """
    errors: list[dict] = []
    imported = 0
    total_rows = 0

    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return {"total_rows": 0, "imported": 0, "errors": [{"row": 0, "error": "تعذر قراءة ملف Excel"}]}

    ws = wb.active
    if ws is None:
        return {"total_rows": 0, "imported": 0, "errors": [{"row": 0, "error": "الملف لا يحتوي على ورقة عمل"}]}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "imported": 0, "errors": [{"row": 0, "error": "الملف فارغ"}]}

    # ── Validate header row ──────────────────────────────────────────────
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    for idx, expected in enumerate(EXPECTED_HEADERS):
        if idx >= len(header) or header[idx] != expected:
            return {
                "total_rows": 0,
                "imported": 0,
                "errors": [{"row": 1, "error": f"عنوان العمود {idx + 1} يجب أن يكون '{expected}' — وجد '{header[idx] if idx < len(header) else ''}'"}],
            }

    # ── Process data rows ────────────────────────────────────────────────
    data_rows = rows[1:]
    total_rows = len(data_rows)

    for row_idx, row in enumerate(data_rows, start=2):
        # Pad row to expected length
        cells = list(row) + [None] * (len(EXPECTED_HEADERS) - len(row))

        prompt = str(cells[0]).strip() if cells[0] is not None else ""
        q_type_raw = str(cells[1]).strip() if cells[1] is not None else ""
        choice_1 = str(cells[2]).strip() if cells[2] is not None else ""
        choice_2 = str(cells[3]).strip() if cells[3] is not None else ""
        choice_3 = str(cells[4]).strip() if cells[4] is not None else ""
        choice_4 = str(cells[5]).strip() if cells[5] is not None else ""
        correct_raw = str(cells[6]).strip() if cells[6] is not None else ""
        score_raw = cells[7]
        difficulty_raw = str(cells[8]).strip() if cells[8] is not None else ""
        category_raw = str(cells[9]).strip() if cells[9] is not None else ""

        # Skip entirely empty rows
        if not prompt and not q_type_raw and not correct_raw:
            total_rows -= 1
            continue

        # ── Validate prompt ──────────────────────────────────────────────
        if not prompt:
            errors.append({"row": row_idx, "error": "نص السؤال فارغ"})
            continue

        # ── Validate question type ───────────────────────────────────────
        if q_type_raw not in VALID_TYPES:
            errors.append({"row": row_idx, "error": f"نوع السؤال غير صالح: '{q_type_raw}' — المسموح: multiple_choice, true_false"})
            continue
        question_type = VALID_TYPES[q_type_raw]

        # ── Build choices list ───────────────────────────────────────────
        choices = [c for c in [choice_1, choice_2, choice_3, choice_4] if c]
        if question_type == QuestionType.MULTIPLE_CHOICE and len(choices) < 2:
            errors.append({"row": row_idx, "error": "أسئلة الاختيار المتعدد تحتاج خيارين على الأقل"})
            continue
        if question_type == QuestionType.TRUE_FALSE and len(choices) < 2:
            errors.append({"row": row_idx, "error": "أسئلة صح/خطأ تحتاج خيارين"})
            continue

        # ── Validate correct answer ──────────────────────────────────────
        if not correct_raw:
            errors.append({"row": row_idx, "error": "الإجابة الصحيحة فارغة"})
            continue
        if correct_raw not in choices:
            errors.append({"row": row_idx, "error": f"الإجابة الصحيحة '{correct_raw}' غير موجودة ضمن الخيارات"})
            continue

        # ── Validate score ───────────────────────────────────────────────
        try:
            score_value = int(score_raw) if score_raw is not None and str(score_raw).strip() else 10
        except (ValueError, TypeError):
            errors.append({"row": row_idx, "error": f"قيمة النقاط غير صالحة: '{score_raw}'"})
            continue
        if score_value <= 0:
            errors.append({"row": row_idx, "error": "النقاط يجب أن تكون أكبر من صفر"})
            continue

        # ── Validate difficulty ───────────────────────────────────────────
        difficulty_key = difficulty_raw if difficulty_raw else "medium"
        if difficulty_key not in VALID_DIFFICULTIES:
            errors.append({"row": row_idx, "error": f"مستوى الصعوبة غير صالح: '{difficulty_key}' — المسموح: easy, medium, hard"})
            continue
        difficulty = VALID_DIFFICULTIES[difficulty_key]

        # ── Category (optional) ──────────────────────────────────────────
        category = category_raw if category_raw else None

        # ── Create question ──────────────────────────────────────────────
        question = Question(
            group_id=group_id,
            question_type=question_type,
            prompt=prompt,
            options={"choices": choices, "correct": correct_raw},
            correct_answer={"answer": correct_raw},
            score_value=score_value,
            difficulty=difficulty,
            category=category,
            status=QuestionStatus.ACTIVE,
        )
        session.add(question)
        imported += 1

    await session.flush()
    wb.close()

    return {
        "total_rows": total_rows,
        "imported": imported,
        "errors": errors,
    }


# ── Export ────────────────────────────────────────────────────────────────


async def export_questions_to_excel(
    session: AsyncSession,
    group_id: uuid.UUID,
) -> bytes:
    """Export all active questions in a group to Excel bytes."""
    result = await session.execute(
        select(Question)
        .where(Question.group_id == group_id, Question.status == QuestionStatus.ACTIVE)
        .order_by(Question.display_order, Question.created_at)
    )
    questions = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "الأسئلة"

    # RTL sheet direction
    ws.sheet_view.rightToLeft = True

    # Write headers
    ws.append(EXPECTED_HEADERS)

    # Write data rows
    for q in questions:
        options = q.options or {}
        choices = options.get("choices", [])
        correct = options.get("correct", q.correct_answer.get("answer", "") if q.correct_answer else "")

        # Pad choices to 4 slots
        padded = choices + [""] * (4 - len(choices))

        ws.append([
            q.prompt,
            q.question_type,
            padded[0],
            padded[1],
            padded[2],
            padded[3],
            correct,
            q.score_value,
            q.difficulty,
            q.category or "",
        ])

    # Auto-adjust column widths for readability
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()
